import pdfplumber
import pandas as pd
import re
from datetime import datetime
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
import glob
import numpy as np
from PyPDF2 import PdfReader, PdfWriter
import io

def get_college_name(pdf_path):
    """
    Extract college name from the BASIC INFORMATION table in the PDF.
    
    Args:
        pdf_path (str): Path to the PDF file
        
    Returns:
        str: College name or None if not found
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if "BASIC INFORMATION" in text:
                    tables = page.extract_tables()
                    for table in tables:
                        # Convert table cells to strings and check if we found the right table
                        str_table = [[str(cell).strip() if cell is not None else "" for cell in row] for row in table]
                        for i, row in enumerate(str_table):
                            # Look for the row with college name (typically second row)
                            if i > 0 and len(row) > 1:  # Skip header row
                                college_name = row[1].strip()
                                if college_name and college_name.lower() != "name of the college":
                                    return college_name
    except Exception as e:
        print(f"Error extracting college name from {pdf_path}: {str(e)}")
    return None

def get_pdf_files(folder_path):
 
    # Use glob to get all PDF files (case insensitive)
    pdf_pattern = os.path.join(folder_path, '**', '*.[pP][dD][fF]')
    pdf_files = glob.glob(pdf_pattern, recursive=True)
    
    if not pdf_files:
        print(f"No PDF files found in: {folder_path}")
        return []
    
    print(f"Found {len(pdf_files)} PDF files")
    return pdf_files

def find_text_and_crop(pdf_path, search_text):
    """
    Find the specified text in PDF and return the page number and y-coordinate
    
    Args:
        pdf_path (str): Path to PDF file
        search_text (str): Text pattern to search for
        
    Returns:
        tuple: (page_number, y_coordinate) or (None, None) if not found
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                text = page.extract_text()
                match = re.search(search_text, text, re.IGNORECASE)
                
                if match:
                    # Find the y-coordinate of the matching text
                    words = page.extract_words()
                    for word in words:
                        if match.group(1) in word['text']:
                            return page_num, word['bottom']
    except Exception as e:
        print(f"Error finding text in {pdf_path}: {str(e)}")
    return None, None

def extract_table_from_cropped_area(pdf_path, page_num, y_coord):
    """
    Extract the first valid table from the cropped area below the specified y-coordinate
    
    Args:
        pdf_path (str): Path to PDF file
        page_num (int): Page number containing the table
        y_coord (float): Y-coordinate to crop from
        
    Returns:
        tuple: (DataFrame, section_number) or (None, None) if no valid table found
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            page = pdf.pages[page_num]
            
            # Crop the page from the y-coordinate to the bottom
            cropped_page = page.crop((0, y_coord, page.width, page.height))
            
            # Extract tables from cropped area
            tables = cropped_page.extract_tables()
            
            # Find the first table with correct structure
            for table in tables:
                if len(table) >= 2 and len(table[0]) == 5:
                    first_row = table[0]
                    if all(re.match(r'\d{4}-\d{2}', str(cell)) for cell in first_row):
                        # Create DataFrame with only first two rows
                        df = pd.DataFrame(table[:2])
                        return df, None
                        
    except Exception as e:
        print(f"Error extracting table from {pdf_path}: {str(e)}")
    return None, None

def extract_enrollment_table(pdf_path):
    """
    Extract enrollment table using text search and page cropping
    """
    # Search pattern
    enrollment_pattern = r'((?:\d+\.)+\d+)\s*Number\s+of\s+students\s+from\s+other\s+states\s+and\s+countries\s+year-wise\s+during\s+the\s+last\s+five\s+years'
    # DONE: r'((?:\d+\.)+\d+)\s*Number\s+of\s+seats\s+filled\s+year\s+wise\s+during\s+last\s+five\s+years'
    # DONE r'((?:\d+\.)+\d+)\s*Number\s+of\s+students\s+admitted\s+year-wise\s+during\s+last\s+five\s+years' 
    # DONE2018__: 2.1.2.1 r'((?:\d+\.)+\d+)\s*Number\s+of\s+seats\s+available\s+year-wise\s+during\s+the\s+last\s+five\s+years'
    # DONE2020+__: 2.1.1.2 r'((?:\d+\.)+\d+)\s*Number\s+of\s+sanctioned\s+seats\s+year\s+wise\s+during\s+last\s+five\s+years'

    # r'((?:\d+\.)+\d+)\s*Amount\s+of\s+seed\s+money\s+provided\s+by\s+institution\s+to\s+its\s+teachers\s+for\s+research\s+year\s+wise\s+during\s+last\s+five\s+years'  
    # r'((?:\d+\.)+\d+)\s*Number\s+of\s+students\s+appeared\s+in\s+the\s+examination\s+conducted\s+by\s+the\s+institution\s+year\s+wise\s+during\s+the\s+last\s+five\s+years'
    # r'((?:\d+\.)+\d+)\s*Number\s+of\s+final\s+year\s+outgoing\s+students\s+year\s+wise\s+during\s+last\s+five\s+years'
    
    # Find text and get coordinates
    page_num, y_coord = find_text_and_crop(pdf_path, enrollment_pattern)
    
    if page_num is not None and y_coord is not None:
        print(f"Found target text in {pdf_path} at page {page_num + 1}, y-coordinate {y_coord}")
        return extract_table_from_cropped_area(pdf_path, page_num, y_coord)
    
    return None, None

def save_to_excel(dataframes, output_path):
    """
    Save the data to Excel with colleges in first row and table data in second row
    """
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        workbook = writer.book
        worksheet = workbook.create_sheet("College Data", 0)
        
        current_col = 1
        default_years = ['x-x', 'x-x', 'x-x', 'x-x', 'x-x']
        
        for pdf_path, (df, _) in dataframes.items():
            college_name = get_college_name(pdf_path) or os.path.splitext(os.path.basename(pdf_path))[0]
            
            # Write college name
            cell = worksheet.cell(row=1, column=current_col)
            cell.value = college_name
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            
            # Merge cells for college name
            worksheet.merge_cells(
                start_row=1,
                start_column=current_col,
                end_row=1,
                end_column=current_col + 4
            )
            
            if df is not None and not df.empty:
                # Write first two rows of the table
                for i, value in enumerate(df.iloc[0]):
                    cell = worksheet.cell(row=2, column=current_col + i)
                    cell.value = value
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.font = Font(color='FFFFFF', bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                if len(df) > 1:  # Write second row if it exists
                    for i, value in enumerate(df.iloc[1]):
                        cell = worksheet.cell(row=3, column=current_col + i)
                        cell.value = value
                        cell.alignment = Alignment(horizontal='center')
            else:
                # Use default values for missing data
                for i, year in enumerate(default_years):
                    cell = worksheet.cell(row=2, column=current_col + i)
                    cell.value = year
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    cell.font = Font(color='FFFFFF', bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    
                    cell = worksheet.cell(row=3, column=current_col + i)
                    cell.value = 'x'
                    cell.alignment = Alignment(horizontal='center')
            
            # Adjust column widths
            for i in range(5):
                col_letter = get_column_letter(current_col + i)
                worksheet.column_dimensions[col_letter].width = 15
            
            current_col += 5
        
        # Remove default sheet if it exists
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])

def create_concatenated_csv(dataframes, all_pdf_files, output_dir):
    """
    Create a concatenated CSV with all college data side by side
    """
    processed_dfs = []
    college_headers = {}
    template_columns = ['2020-21', '2019-20', '2018-19', '2017-18', '2016-17']
    
    for pdf_path in all_pdf_files:
        college_name = get_college_name(pdf_path) or os.path.splitext(os.path.basename(pdf_path))[0]
        
        if pdf_path in dataframes and dataframes[pdf_path][0] is not None:
            df = dataframes[pdf_path][0]
            # Use only first two rows
            college_df = df.iloc[:2]
            
            for col in df.columns:
                college_headers[col] = college_name
        else:
            empty_data = {year: [np.nan, np.nan] for year in template_columns}
            college_df = pd.DataFrame(empty_data)
            
            for col in empty_data.keys():
                college_headers[col] = college_name
        
        processed_dfs.append(college_df)
    
    if processed_dfs:
        final_df = pd.concat(processed_dfs, axis=1)
        college_names_row = pd.DataFrame([college_headers], columns=final_df.columns)
        final_df = pd.concat([college_names_row, final_df], ignore_index=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_path = os.path.join(output_dir, f'criteria_{timestamp}.csv')
        final_df.to_csv(csv_path, index=False)
        return csv_path
    return None

def process_folder(input_folder, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    pdf_files = get_pdf_files(input_folder)
    
    if not pdf_files:
        return
    
    results = {}
    for pdf_path in pdf_files:
        try:
            print(f"Processing: {pdf_path}")
            df, section_number = extract_enrollment_table(pdf_path)
            results[pdf_path] = (df, section_number)
            
            if df is not None:
                print(f"✓ Successfully extracted table from: {os.path.basename(pdf_path)}")
            else:
                print(f"✗ No matching table found in: {os.path.basename(pdf_path)}")
        except Exception as e:
            print(f"Error processing {pdf_path}: {str(e)}")
            results[pdf_path] = (None, None)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(output_dir, f'enrollment_data_{timestamp}.xlsx')
    save_to_excel(results, excel_path)
    print(f"Excel output saved to: {excel_path}")
    
  
    
    print(f"\nProcessing Summary:")
    print(f"Total PDFs processed: {len(pdf_files)}")
    successful_extractions = sum(1 for _, (df, _) in results.items() if df is not None)
    print(f"Successfully extracted tables: {successful_extractions}")
    print(f"Failed extractions: {len(pdf_files) - successful_extractions}")
    
    

if __name__ == "__main__":
    input_folder = "Atotally"  # Folder containing PDFs
    output_dir = "result/twopoints"  # Where to save the results
    process_folder(input_folder, output_dir)